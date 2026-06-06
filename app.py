#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
تطبيق معالجة صور وبيانات الطلبة الخريجين
جمع صور الطلبة من مايكروسوفت فورم وتنظيمها حسب الصف والمهنة
"""

import openpyxl, os, re, shutil, sys, json, time
from urllib.parse import unquote
from collections import defaultdict
from datetime import datetime
import winreg

# ======================== CONFIG ========================
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')
DEFAULT_CONFIG = {
    "excel_folder": "",  # Auto: Desktop folder named "استمارات"
    "output_folder": "",  # Auto: same as script folder
    "auto_detect_onedrive": True,
}

def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return dict(DEFAULT_CONFIG)

def save_config(config):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

# ======================== DETECT PATHS ========================
def find_onedrive_business():
    """Find OneDrive for Business sync folder from registry"""
    try:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'Software\Microsoft\OneDrive\Accounts\Business1')
        path = winreg.QueryValueEx(key, 'UserFolder')[0]
        if os.path.exists(path):
            return path
    except:
        pass
    return None

def find_form_folder(onedrive_root):
    """Find the Microsoft Forms folder containing graduation images"""
    forms_root = os.path.join(onedrive_root, 'Apps', 'Microsoft Forms')
    if not os.path.exists(forms_root):
        return None
    
    # Look for the most recent form folder with images
    best = None
    best_time = 0
    for item in os.listdir(forms_root):
        full = os.path.join(forms_root, item)
        if os.path.isdir(full):
            q = os.path.join(full, 'Question')
            if os.path.isdir(q):
                files = os.listdir(q)
                if files:
                    # Check newest file time
                    for f in files:
                        fp = os.path.join(q, f)
                        t = os.path.getmtime(fp)
                        if t > best_time:
                            best_time = t
                            best = (item, q)
    
    return best

def find_latest_excel(folder):
    """Find the latest Excel export from Microsoft Forms"""
    if not folder or not os.path.exists(folder):
        return None
    excel_files = []
    for f in os.listdir(folder):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            full = os.path.join(folder, f)
            excel_files.append((os.path.getmtime(full), full, f))
    
    if not excel_files:
        return None
    excel_files.sort(reverse=True)
    return excel_files[0]

def find_question_folder(onedrive_root, form_name):
    """Find the Question folder for a specific form name"""
    forms_root = os.path.join(onedrive_root, 'Apps', 'Microsoft Forms')
    if not os.path.exists(forms_root):
        return None
    # Search for the form folder (name may differ slightly)
    for item in os.listdir(forms_root):
        # Match by common keywords
        if ('سادس' in item or 'خريج' in item):
            q = os.path.join(forms_root, item, 'Question')
            if os.path.isdir(q):
                return q
    return None

# ======================== PROFESSION NORMALIZATION ========================
def normalize_profession(prof):
    prof = prof.strip()
    # مهندس بدون تفصيل → مهندس معماري
    if prof == 'مهندس' or prof == 'مهندس ':
        return 'مهندس معماري'
    return prof

def safe_filename(text):
    """Remove characters that are invalid in Windows filenames"""
    invalid = r'\/:*?"<>|'
    for c in invalid:
        text = text.replace(c, '_')
    return text.strip()

# ======================== MAIN PROCESSING ========================
def process(onedrive_root, excel_path, output_root, progress_callback=None):
    """Main processing function"""
    
    def log(msg):
        if progress_callback:
            progress_callback(msg)
        else:
            print(msg)
    
    # ---- Load Excel ----
    log(f"📂 قراءة ملف Excel: {os.path.basename(excel_path)}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    students = []
    for row in range(2, ws.max_row + 1):
        agree = ws.cell(row, 7).value
        if agree != 'موافق':
            continue
        
        cls = ws.cell(row, 8).value
        
        arabic_name = None
        for col in range(9, 14):
            v = ws.cell(row, col).value
            if v:
                arabic_name = str(v).strip()
                break
        
        eng_name = str(ws.cell(row, 5).value or '').strip()
        image_url = str(ws.cell(row, 14).value or '')
        profession = str(ws.cell(row, 15).value or '').strip()
        
        filename_from_url = unquote(image_url.split('/')[-1]) if image_url else ''
        
        students.append({
            'row': row,
            'class': cls,
            'arabic_name': arabic_name,
            'eng_name': eng_name,
            'profession': profession,
            'url_filename': filename_from_url,
        })
    
    log(f"✅ عدد الطلبة في الاستمارة: {len(students)}")
    
    # ---- Find form Question folder ----
    question_folder = find_question_folder(onedrive_root, None)
    if not question_folder:
        log("❌ لم يتم العثور على مجلد صور الاستمارة في OneDrive")
        log("   تأكد من مزامنة OneDrive للمجلد: Apps/Microsoft Forms/...")
        return False
    
    log(f"📁 مجلد الصور: {question_folder}")
    
    # ---- Load source images ----
    source_files = {}
    for f in os.listdir(question_folder):
        full = os.path.join(question_folder, f)
        if os.path.isfile(full):
            source_files[f] = full
    
    log(f"🖼️ عدد ملفات الصور الموجودة: {len(source_files)}")
    
    # ---- Match students to images ----
    matched = 0
    unmatched_students = 0
    
    for s in students:
        url_fn = s['url_filename']
        ext = '.jpg'
        
        if url_fn in source_files:
            s['source_file'] = source_files[url_fn]
            ext = os.path.splitext(url_fn)[1]
            s['ext'] = ext
            matched += 1
            continue
        
        # Fuzzy match by English name
        best_match = None
        best_score = 0
        s_eng = s['eng_name'].lower()
        
        for local_fn in source_files:
            l_fn_lower = local_fn.lower()
            name_words = s_eng.split()
            matched_words = sum(1 for w in name_words if len(w) > 2 and w in l_fn_lower)
            score = matched_words / max(len(name_words), 1)
            if score > best_score:
                best_score = score
                best_match = local_fn
        
        if best_match and best_score >= 0.3:
            s['source_file'] = source_files[best_match]
            s['ext'] = os.path.splitext(best_match)[1]
            matched += 1
        else:
            s['source_file'] = None
            unmatched_students += 1
            log(f"⚠ لم يتم العثور على صورة: {s['arabic_name']} ({s['url_filename']})")
    
    log(f"✅ تمت مطابقة {matched} من أصل {len(students)} طالب")
    if unmatched_students:
        log(f"⚠ {unmatched_students} طالب/طالبة بدون صور (قد تكون الصورة لم ترفع بعد)")
    
    # ---- Create output folders ----
    by_class = os.path.join(output_root, 'حسب الصف')
    by_profession = os.path.join(output_root, 'حسب المهنة')
    by_both = os.path.join(output_root, 'حسب الصف والمهنة')
    
    for folder in [by_class, by_profession, by_both]:
        os.makedirs(folder, exist_ok=True)
    
    # ---- Process ----
    name_counter = defaultdict(int)
    copied_count = 0
    skipped_count = 0
    
    for s in students:
        if 'source_file' not in s or not s['source_file']:
            continue
        
        cls = s['class']
        arabic = s['arabic_name']
        prof_raw = s['profession']
        prof_norm = normalize_profession(prof_raw)
        ext = s['ext']
        src = s['source_file']
        
        safe_name = safe_filename(arabic)
        safe_prof = safe_filename(prof_norm)
        
        key = f"{arabic}|{cls}"
        name_counter[key] += 1
        count = name_counter[key]
        
        if count > 1:
            new_name = f"{safe_name} ({count}) - {cls} - {safe_prof}{ext}"
        else:
            new_name = f"{safe_name} - {cls} - {safe_prof}{ext}"
        
        # To each folder
        for base_folder in [by_class, by_profession, by_both]:
            if base_folder == by_class:
                sub = os.path.join(base_folder, cls)
            elif base_folder == by_profession:
                sub = os.path.join(base_folder, prof_norm)
            else:
                sub = os.path.join(base_folder, cls, prof_norm)
            
            os.makedirs(sub, exist_ok=True)
            dest = os.path.join(sub, new_name)
            if not os.path.exists(dest):
                shutil.copy2(src, dest)
                copied_count += 1
            else:
                skipped_count += 1
    
    log(f"✅ تم نسخ {copied_count} صورة (تخطي {skipped_count} صورة موجودة مسبقاً)")
    
    # ---- Print summary ----
    log(f"\n{'='*50}")
    log(f"📊 ملخص المعالجة:")
    log(f"{'='*50}")
    log(f"📂 حسب الصف:")
    for c in sorted(os.listdir(by_class)):
        count = len(os.listdir(os.path.join(by_class, c)))
        log(f"   {c}: {count} طالب")
    log(f"📂 حسب المهنة:")
    for p in sorted(os.listdir(by_profession)):
        count = len(os.listdir(os.path.join(by_profession, p)))
        log(f"   {p}: {count} طالب")
    
    log(f"\n📁 تم حفظ الملفات في: {output_root}")
    
    return True

# ======================== INTERACTIVE MODE ========================
def run_interactive():
    print("=" * 60)
    print("           📋 تطبيق تنظيم صور الطلبة الخريجين")
    print("=" * 60)
    print()
    
    config = load_config()
    
    # ---- Detect OneDrive ----
    print("🔍 جاري البحث عن مجلد OneDrive للمؤسسة التعليمية...")
    onedrive_root = find_onedrive_business()
    if onedrive_root:
        print(f"✅ تم العثور على OneDrive: {onedrive_root}")
    else:
        print("❌ لم يتم العثور على OneDrive للمؤسسة التعليمية")
        path = input("   الرجاء إدخال المسار يدوياً: ").strip()
        if os.path.exists(path):
            onedrive_root = path
        else:
            print("❌ المسار غير صحيح")
            input("\nاضغط Enter للخروج...")
            return
    
    # ---- Get Excel file ----
    print("\n🔍 جاري البحث عن ملف Excel المستورد من Microsoft Forms...")
    
    excel_path = None
    
    # Option 1: Use configured folder
    if config.get('excel_folder') and os.path.exists(config['excel_folder']):
        result = find_latest_excel(config['excel_folder'])
        if result:
            excel_path = result[1]
            print(f"✅ تم العثور على ملف: {result[2]}")
    
    # Option 2: Search Desktop
    if not excel_path:
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        for item in os.listdir(desktop):
            # Look for Excel files with related names
            if item.endswith('.xlsx') and not item.startswith('~$'):
                if 'سادس' in item or 'خريج' in item or 'امنيات' in item or 'أمنيات' in item:
                    candidate = os.path.join(desktop, item)
                    excel_path = candidate
                    print(f"✅ تم العثور على ملف على سطح المكتب: {item}")
                    break
    
    # Option 3: Ask user
    if not excel_path or not os.path.exists(excel_path):
        print("📂 يرجى اختيار ملف Excel المستورد من Microsoft Forms:")
        from tkinter import filedialog, Tk
        root = Tk()
        root.withdraw()
        excel_path = filedialog.askopenfilename(
            title="اختر ملف Excel",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        root.destroy()
        
        if not excel_path:
            print("❌ لم يتم اختيار ملف")
            input("\nاضغط Enter للخروج...")
            return
    
    print(f"📂 ملف Excel: {excel_path}")
    
    # ---- Output folder ----
    output_root = config.get('output_folder', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output'))
    if not output_root:
        output_root = os.path.join(os.path.dirname(excel_path), 'output_طلبة')
    
    os.makedirs(output_root, exist_ok=True)
    print(f"📁 مجلد الإخراج: {output_root}")
    
    # ---- Confirm ----
    print("\n" + "=" * 50)
    print("🚀 جاري معالجة البيانات والصور...")
    print("=" * 50)
    print()
    
    # ---- Process ----
    success = process(onedrive_root, excel_path, output_root)
    
    if success:
        print(f"\n{'='*50}")
        print("✅ تمت المعالجة بنجاح!")
        print(f"📁 الملفات في: {output_root}")
        print(f"{'='*50}")
    
    # ---- Save config ----
    config['excel_folder'] = os.path.dirname(excel_path)
    config['output_folder'] = output_root
    save_config(config)
    
    print()
    input("👆 اضغط Enter للخروج...")

def run_quick(excel_path=None, output_root=None):
    """Quick mode - run without interaction"""
    config = load_config()
    
    onedrive_root = find_onedrive_business()
    if not onedrive_root:
        print("❌ OneDrive للمؤسسة التعليمية غير موجود")
        return False
    
    if not excel_path:
        folder = config.get('excel_folder', os.path.join(os.path.expanduser('~'), 'Desktop'))
        result = find_latest_excel(folder)
        if not result:
            print("❌ لم يتم العثور على ملف Excel")
            return False
        excel_path = result[1]
        print(f"📂 استخدام: {result[2]}")
    
    if not output_root:
        output_root = config.get('output_folder', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output'))
    
    os.makedirs(output_root, exist_ok=True)
    
    return process(onedrive_root, excel_path, output_root)

# ======================== ENTRY POINT ========================
if __name__ == '__main__':
    # Check for command line arguments for quick mode
    if len(sys.argv) > 1:
        excel = sys.argv[1] if len(sys.argv) > 1 else None
        output = sys.argv[2] if len(sys.argv) > 2 else None
        success = run_quick(excel, output)
        if success:
            print("\n✅ تم بنجاح!")
        else:
            print("\n❌ فشلت المعالجة")
            sys.exit(1)
    else:
        try:
            run_interactive()
        except KeyboardInterrupt:
            print("\n\n❌ تم إلغاء العملية")
            sys.exit(1)
        except Exception as e:
            print(f"\n❌ خطأ: {e}")
            import traceback
            traceback.print_exc()
            input("\nاضغط Enter للخروج...")
            sys.exit(1)
